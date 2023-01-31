import requests
import shutil
import urllib
from openpyxl import load_workbook

wb = load_workbook(filename='BazaPytanS.xlsx')
sheet = wb['SOPER']

for x in range(489):

    text = 'H{}'.format(x+1)
    if sheet[text].value:
        print(x,sheet[text].value)
        image_url = sheet[text].value
        filename = "SOPER{}.png".format(x+1)
        def download_jpg(url, file_path, file_name):
            full_path = file_path + file_name
            urllib.request.urlretrieve(url, full_path)
        download_jpg(image_url, r"D:\Project\Image\Systemy/", filename)
    else:
        print(x, "https://egzamin-informatyk.pl/assets/images/sprzet-systemy-ee08-inf02.jpg")
        filename = "SOPER{}.png".format(x+1)
        def download_jpg(url, file_path, file_name):
            full_path = file_path + file_name
            urllib.request.urlretrieve(url, full_path)
        download_jpg("https://egzamin-informatyk.pl/assets/images/sprzet-systemy-ee08-inf02.jpg", r"D:\Project\Image\Systemy/", filename)


wb = load_workbook(filename='BazaPytanU.xlsx')
sheet = wb['UTK']
for x in range(694):

    text = 'H{}'.format(x+1)
    if sheet[text].value:
        print(x,sheet[text].value)
        image_url = sheet[text].value
        filename = "UTK{}.png".format(x+1)
        def download_jpg(url, file_path, file_name):
            full_path = file_path + file_name
            urllib.request.urlretrieve(url, full_path)
        download_jpg(image_url, r"D:\Project\Image\Urzadzenia/", filename)
    else:
        print(x, "https://egzamin-informatyk.pl/assets/images/sprzet-systemy-ee08-inf02.jpg")
        filename = "UTK{}.png".format(x+1)
        def download_jpg(url, file_path, file_name):
            full_path = file_path + file_name
            urllib.request.urlretrieve(url, full_path)
        download_jpg("https://egzamin-informatyk.pl/assets/images/sprzet-systemy-ee08-inf02.jpg", r"D:\Project\Image\Urzadzenia/", filename)

wb = load_workbook(filename='BazaPytanE.xlsx')
sheet = wb['ELSK']
for x in range(1090):

    text = 'H{}'.format(x+1)
    if sheet[text].value:
        print(x,sheet[text].value)
        image_url = sheet[text].value
        filename = "ELSK{}.png".format(x+1)
        def download_jpg(url, file_path, file_name):
            full_path = file_path + file_name
            urllib.request.urlretrieve(url, full_path)
        download_jpg(image_url, r"D:\Project\Image\Sieci/", filename)
    else:
        print(x, "https://egzamin-informatyk.pl/assets/images/sprzet-systemy-ee08-inf02.jpg")
        filename = "ELSK{}.png".format(x+1)
        def download_jpg(url, file_path, file_name):
            full_path = file_path + file_name
            urllib.request.urlretrieve(url, full_path)
        download_jpg("https://egzamin-informatyk.pl/assets/images/sprzet-systemy-ee08-inf02.jpg", r"D:\Project\Image\Sieci/", filename)
